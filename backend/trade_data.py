"""Trade-workbook parsing + metrics (extracted from main.py).

Pure functions the analytics layer is built on:

- `read_trades_excel` — read a trades .xlsx into a DataFrame, RE-EVALUATING the
  literal-arithmetic formula cells the formatter writes for scale-out (multi-fill)
  exits. pandas reads those as NaN, which would silently drop every scaled-out
  trade — the single most important quirk in the pipeline (see tests/test_trade_data.py).
- `normalize_trade_data` — DataFrame -> list of *closed*-trade dicts (rows with no
  exit price are dropped) with typed / cleaned fields.
- `calculate_trade_metrics` — trade dicts -> win rate, profit factor, averages.

No dependency on main.py: the in-memory mtime cache and the review-notes overlay
stay there. Only `read_trades_excel` touches the filesystem (opening the workbook).
"""

import re

import numpy as np
import pandas as pd


_ARITH_RE = re.compile(r'^[0-9eE+\-*/().\s]+$')

# Numeric columns whose cells can hold partial-fill weighted-average formulas
# (e.g. a scale-out exit written as =((100*49.401)+(125*54.61))/225). pandas
# reads those formula cells as NaN, which silently drops the trade from the
# analysis — so we evaluate the literal arithmetic on read.
_FORMULA_NUMERIC_COLS = {
    'Qty', 'Entry Price', 'Exit Qty', 'Exit Price', 'Stop Price', 'Target Price',
}


def _eval_arith_formula(expr):
    """Safely evaluate a pure-arithmetic Excel formula like '=18+2' or
    '=((132.2*18)+(2*108.88))/20'. Returns a float, or None if the string
    isn't a self-contained arithmetic expression (e.g. it references cells)."""
    if not isinstance(expr, str) or not expr.startswith('='):
        return None
    s = expr[1:].replace(',', '').strip()
    if not s or not _ARITH_RE.match(s):
        return None
    try:
        return float(eval(s, {"__builtins__": {}}, {}))  # noqa: S307 - regex-guarded, no names
    except Exception:
        return None


def read_trades_excel(source):
    """Read a trades workbook into a DataFrame, evaluating literal-arithmetic
    formula cells that pandas returns as NaN. Without this, every multi-fill
    (scale-out) exit — stored by the formatter as a weighted-average formula —
    is dropped from the analytics. Falls back to a plain read on any error."""
    df = pd.read_excel(source)
    try:
        from openpyxl import load_workbook
        if hasattr(source, 'seek'):
            source.seek(0)
        wb = load_workbook(source, data_only=False, read_only=True)
        ws = wb['Trades'] if 'Trades' in wb.sheetnames else wb.active
        col_pos = {}
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if ri == 0:
                for i, name in enumerate(row):
                    if name in _FORMULA_NUMERIC_COLS and name in df.columns:
                        col_pos[name] = (i, df.columns.get_loc(name))
                continue
            dfi = ri - 1
            if dfi >= len(df):
                break
            for name, (xi, di) in col_pos.items():
                if pd.isna(df.iat[dfi, di]):
                    val = _eval_arith_formula(row[xi] if xi < len(row) else None)
                    if val is not None:
                        df.iat[dfi, di] = val
        wb.close()
    except Exception:
        pass
    return df


def normalize_trade_data(df):
    """
    Normalize trade data from various formats to a consistent structure.
    Handles the specific format from Trades_2025.xlsx with columns like:
    Symbol, Qty, Side, Entry Price, Entry Date, Exit Price, Exit Date, Profit / Loss, Profit / Loss %
    """
    # Column mapping - handles different naming conventions
    column_mapping = {
        'Symbol': 'symbol',
        'Qty': 'quantity',
        'Side': 'side',
        'Entry Price': 'entry_price',
        'Entry Date': 'entry_date',
        'Entry Time': 'entry_time',
        'Exit Price': 'exit_price',
        'Exit Date': 'exit_date',
        'Exit Time': 'exit_time',
        'Exit Qty': 'exit_quantity',
        'Profit / Loss': 'pnl',
        'Profit / Loss %': 'pnl_pct',
        'Setup': 'setup',
        'Entry Notes': 'entry_notes',
        'Notes': 'notes',
        'Market Cap': 'market_cap',
        'Stock/Option': 'instrument_type',
        'Emotion': 'emotion',
        'Conviction': 'conviction',
        'Stop Price': 'stop_price',
        'Target Price': 'target_price',
        'Grade': 'grade',
    }

    # Rename columns if they exist
    df_renamed = df.rename(columns=column_mapping)

    # Filter only closed trades (those with exit prices)
    df_filtered = df_renamed[df_renamed['exit_price'].notna()].copy()

    # Ensure date columns are datetime
    if 'entry_date' in df_filtered.columns:
        df_filtered['entry_date'] = pd.to_datetime(df_filtered['entry_date'], errors='coerce')
    if 'exit_date' in df_filtered.columns:
        df_filtered['exit_date'] = pd.to_datetime(df_filtered['exit_date'], errors='coerce')

    # Calculate trade duration in days
    if 'entry_date' in df_filtered.columns and 'exit_date' in df_filtered.columns:
        df_filtered['duration_days'] = (df_filtered['exit_date'] - df_filtered['entry_date']).dt.days

    # P&L — filled PER ROW, never all-or-nothing.
    #
    # The workbook stores P/L as an Excel formula, `=(Exit Qty * Exit Price) -
    # (Qty * Entry Price)`, and P/L % as `=P/(Qty * Entry Price)` (a decimal
    # fraction). Those are cell-referencing formulas, so `read_trades_excel`
    # can't evaluate them — it only handles self-contained arithmetic — and any
    # openpyxl save (every formatter run) strips the cached results Excel left
    # behind. pandas then reads the whole column as NaN.
    #
    # This used to be guarded by `isna().all()`, which a single manually-typed
    # number defeats: one recovered row out of 500 kept the guard False, so the
    # other 499 skipped the fallback and were zeroed by the NaN cleanup below —
    # a workbook of real trades reporting $0 P&L and a 0.2% win rate. Recompute
    # the workbook's own expression row-wise and use it wherever there is no
    # cached number, so a partially-cached column heals row by row.
    qty = pd.to_numeric(df_filtered.get('quantity'), errors='coerce')
    entry = pd.to_numeric(df_filtered.get('entry_price'), errors='coerce')
    exit_px = pd.to_numeric(df_filtered.get('exit_price'), errors='coerce')
    # Scale-outs book only the shares actually sold; assume a full exit when the
    # column is absent or blank.
    sold = pd.to_numeric(df_filtered['exit_quantity'], errors='coerce').fillna(qty) \
        if 'exit_quantity' in df_filtered.columns else qty

    computed_pnl = (exit_px * sold) - (entry * qty)
    if 'side' in df_filtered.columns:
        # The sheet's formula is long-biased — a covered short prints its P&L
        # with the sign flipped. Shorts realise the inverse.
        is_short = df_filtered['side'].astype(str).str.strip().str.upper().str.startswith('SHORT')
        computed_pnl = computed_pnl.mask(is_short, (entry * qty) - (exit_px * sold))

    if 'pnl' in df_filtered.columns:
        df_filtered['pnl'] = pd.to_numeric(df_filtered['pnl'], errors='coerce').fillna(computed_pnl)
    else:
        df_filtered['pnl'] = computed_pnl

    # P&L % against the cost basis (matches the workbook's own definition, and
    # unlike a raw price return it stays honest on partial exits).
    cost_basis = (entry * qty).replace(0, np.nan)
    computed_pct = (df_filtered['pnl'] / cost_basis) * 100
    if 'pnl_pct' in df_filtered.columns:
        pct = pd.to_numeric(df_filtered['pnl_pct'], errors='coerce')
        # Stored as a decimal fraction (0.12) rather than a percentage (12)?
        sample_values = pct.dropna().head(10)
        if len(sample_values) > 0 and sample_values.abs().max() < 10:
            pct = pct * 100
        df_filtered['pnl_pct'] = pct.fillna(computed_pct)
    else:
        df_filtered['pnl_pct'] = computed_pct

    # Convert to list of dicts, handling NaN values
    trades = df_filtered.to_dict('records')

    # Clean up NaN values and convert dates to ISO strings.
    #
    # NOTE: a missing value comes through as NaN, which is a *float* — so a
    # naive `isinstance(value, (float, int))` check treats EVERY missing field
    # (including text columns) as numeric and sets it to 0. That made empty
    # setup/grade/etc. render as a stray "0" in the UI (e.g. "BRAI0",
    # "-14.36%0"). Classify columns by name instead.
    _DATE_COLS = {'entry_date', 'exit_date', 'entry_time', 'exit_time'}
    _NULL_NUM_COLS = {'stop_price', 'target_price', 'conviction'}
    _STRING_COLS = {
        'symbol', 'side', 'setup', 'entry_notes', 'notes',
        'instrument_type', 'emotion', 'grade', 'market_cap',
    }
    for trade in trades:
        for key, value in trade.items():
            if pd.isna(value):
                if key in _DATE_COLS:
                    trade[key] = None
                elif key in _NULL_NUM_COLS:
                    trade[key] = None
                elif key in _STRING_COLS:
                    trade[key] = ''
                else:
                    trade[key] = 0
            elif key in ('entry_date', 'exit_date') and hasattr(value, 'isoformat'):
                trade[key] = value.isoformat()

    return trades


def calculate_trade_metrics(trades):
    """Calculate comprehensive trading metrics from trade data."""
    if not trades:
        return {
            "total_pnl": 0,
            "win_rate": 0,
            "avg_win": 0,
            "avg_loss": 0,
            "profit_factor": 0,
            "total_trades": 0,
        }

    df = pd.DataFrame(trades)

    winning_trades = [t for t in trades if t.get('pnl', 0) > 0]
    losing_trades = [t for t in trades if t.get('pnl', 0) <= 0]

    total_pnl = df['pnl'].sum() if 'pnl' in df.columns else 0
    total_trades = len(trades)
    win_count = len(winning_trades)
    loss_count = len(losing_trades)
    win_rate = (win_count / total_trades * 100) if total_trades > 0 else 0

    avg_win = np.mean([t['pnl'] for t in winning_trades]) if winning_trades else 0
    avg_loss = np.mean([t['pnl'] for t in losing_trades]) if losing_trades else 0

    gross_profit = sum([t['pnl'] for t in winning_trades])
    gross_loss = abs(sum([t['pnl'] for t in losing_trades]))
    profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else 0

    return {
        "total_pnl": round(total_pnl, 2),
        "win_rate": round(win_rate, 1),
        "avg_win": round(avg_win, 2),
        "avg_loss": round(avg_loss, 2),
        "profit_factor": round(profit_factor, 2),
        "total_trades": total_trades,
        "winning_trades": win_count,
        "losing_trades": loss_count,
    }
